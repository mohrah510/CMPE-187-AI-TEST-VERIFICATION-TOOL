import sys
sys.dont_write_bytecode = True
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

def save_results(name, results):
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
    output_dir = os.path.join(project_root, "output")
    try:
        os.makedirs(output_dir, exist_ok=True, mode=0o755)
    except PermissionError as e:
        raise PermissionError(f"Cannot create output directory: {str(e)}")
    
    filename = os.path.join(output_dir, f"{name}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    
    row = 1
    
    for idx, r in enumerate(results, start=1):
        
        question = r.get("question", "")
        input_1 = r.get("input_1", "")
        input_2 = r.get("input_2", "")
        input_3 = r.get("input_3", "")
        context_1 = r.get("context_1", "")
        context_2 = r.get("context_2", "")
        context_3 = r.get("context_3", "")
        expected_valid = r.get("expected_valid", "")
        chatgpt_answer = r.get("chatgpt_answer", "")
        deepseek_answer = r.get("deepseek_answer", "")
        chatgpt_score = r.get("chatgpt_score", "")
        deepseek_score = r.get("deepseek_score", "")
        chatgpt_validity = r.get("chatgpt_validity", "")
        deepseek_validity = r.get("deepseek_validity", "")
        
        if name == "airline_policy":
            input_label_1 = "Emirates Policy Inquiry"
            input_label_2 = "Baggage and Cabin"
            input_label_3 = "Question Type"
            context_label_1 = "Special Conditions"
            context_label_2 = "Traveler Profile"
            context_label_3 = "Trip Characteristics"
        else:
            input_label_1 = "UAE Visa Topic"
            input_label_2 = "UAE Visa Types"
            input_label_3 = "UAE Visa Question Type"
            context_label_1 = "Traveler Profile"
            context_label_2 = "Travel Purpose"
            context_label_3 = "Stay Details"
        
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = f"Test Case ID {idx}"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        row += 1
        
        ws[f'A{row}'] = "Input Feature"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'] = ""
        row += 1
        
        ws[f'A{row}'] = input_label_1
        ws[f'B{row}'] = input_1
        row += 1
        
        ws[f'A{row}'] = input_label_2
        ws[f'B{row}'] = input_2
        row += 1
        
        ws[f'A{row}'] = input_label_3
        ws[f'B{row}'] = input_3
        row += 1
        
        row += 1
        
        ws[f'A{row}'] = "Context Feature"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'] = ""
        row += 1
        
        ws[f'A{row}'] = context_label_1
        ws[f'B{row}'] = context_1
        row += 1
        
        ws[f'A{row}'] = context_label_2
        ws[f'B{row}'] = context_2
        row += 1
        
        ws[f'A{row}'] = context_label_3
        ws[f'B{row}'] = context_3
        row += 1
        
        row += 1
        
        ws[f'A{row}'] = "Expected Output"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'] = ""
        row += 1
        
        ws[f'A{row}'] = "Inaccurate Information"
        ws[f'B{row}'] = "Invalid"
        row += 1
        
        ws[f'A{row}'] = "Correct Information"
        ws[f'B{row}'] = "Valid"
        row += 1
        
        row += 1
        
        ws[f'A{row}'] = "Text Input (Question to LLMs)"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'] = question
        ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
        
        ws[f'A{row}'] = "Expected Answer (Ground Truth)"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'] = expected_valid
        ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
        
        ws[f'A{row}'] = "Actual Output (AI Model Responses)"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'] = "ChatGPT"
        ws[f'B{row}'].font = header_font
        ws[f'C{row}'] = "DeepSeek"
        ws[f'C{row}'].font = header_font
        row += 1
        
        # Response preview (truncated, like sample group)
        chatgpt_preview = r.get("chatgpt_preview", chatgpt_answer)
        deepseek_preview = r.get("deepseek_preview", deepseek_answer)
        
        ws[f'A{row}'] = "Answer (Preview)"
        preview_text_gpt = chatgpt_preview + ("..." if len(chatgpt_answer) > len(chatgpt_preview) else "")
        preview_text_ds = deepseek_preview + ("..." if len(deepseek_answer) > len(deepseek_preview) else "")
        ws[f'B{row}'] = preview_text_gpt
        ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f'C{row}'] = preview_text_ds
        ws[f'C{row}'].alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
        
        # Full answer (if different from preview)
        if len(chatgpt_answer) > len(chatgpt_preview) or len(deepseek_answer) > len(deepseek_preview):
            ws[f'A{row}'] = "Answer (Full)"
            ws[f'B{row}'] = chatgpt_answer
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical="top")
            ws[f'C{row}'] = deepseek_answer
            ws[f'C{row}'].alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        
        # Expected outputs with per-keyword status (like sample group)
        expected_keywords = r.get("expected_keywords", [])
        chatgpt_per_keyword = r.get("chatgpt_per_keyword", [])
        deepseek_per_keyword = r.get("deepseek_per_keyword", [])
        
        if expected_keywords:
            ws[f'A{row}'] = "Expected Outputs (Per-Keyword Status)"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'] = "ChatGPT"
            ws[f'B{row}'].font = header_font
            ws[f'C{row}'] = "DeepSeek"
            ws[f'C{row}'].font = header_font
            row += 1
            
            for i, keyword in enumerate(expected_keywords):
                gpt_status = "True" if i < len(chatgpt_per_keyword) and chatgpt_per_keyword[i] else "False"
                ds_status = "True" if i < len(deepseek_per_keyword) and deepseek_per_keyword[i] else "False"
                ws[f'A{row}'] = f"[{keyword}]"
                ws[f'B{row}'] = gpt_status
                ws[f'C{row}'] = ds_status
                row += 1
        
        # Total correct count and percentage (like sample group)
        chatgpt_matched = sum(1 for status in chatgpt_per_keyword if status) if chatgpt_per_keyword else 0
        deepseek_matched = sum(1 for status in deepseek_per_keyword if status) if deepseek_per_keyword else 0
        total_keywords = len(expected_keywords) if expected_keywords else 0
        
        if total_keywords > 0:
            chatgpt_pct = (chatgpt_matched / total_keywords * 100) if total_keywords > 0 else 0
            deepseek_pct = (deepseek_matched / total_keywords * 100) if total_keywords > 0 else 0
            ws[f'A{row}'] = "Total Correct"
            ws[f'B{row}'] = f"{chatgpt_matched} - {chatgpt_pct:.1f}%"
            ws[f'C{row}'] = f"{deepseek_matched} - {deepseek_pct:.1f}%"
            row += 1
        
        ws[f'A{row}'] = "Result (Evaluation of Actual Output)"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'] = "ChatGPT"
        ws[f'B{row}'].font = header_font
        ws[f'C{row}'] = "DeepSeek"
        ws[f'C{row}'].font = header_font
        row += 1
        
        ws[f'A{row}'] = "Validity"
        ws[f'B{row}'] = chatgpt_validity
        ws[f'C{row}'] = deepseek_validity
        row += 1
        
        # Validity reason (like sample group)
        chatgpt_reason = r.get("chatgpt_validity_reason", "")
        deepseek_reason = r.get("deepseek_validity_reason", "")
        if chatgpt_reason or deepseek_reason:
            ws[f'A{row}'] = "Reason"
            ws[f'B{row}'] = chatgpt_reason
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical="top")
            ws[f'C{row}'] = deepseek_reason
            ws[f'C{row}'].alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        
        row += 2
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50

    try:
        wb.save(filename)
        os.chmod(filename, 0o644)  # Make file readable
    except PermissionError as e:
        raise PermissionError(f"Cannot save Excel file '{filename}': Permission denied. Make sure the file is not open in another program.")
    except Exception as e:
        raise Exception(f"Error saving Excel file '{filename}': {str(e)}")


def save_summary(airline_results, visa_results):
    """Create a summary Excel file with pass rates for both test sets"""
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
    output_dir = os.path.join(project_root, "output")
    os.makedirs(output_dir, exist_ok=True)
    
    filename = os.path.join(output_dir, "test_summary.xlsx")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    
    # Styles
    header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  # Light purple
    header_font = Font(bold=True, size=12)
    data_fill_even = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Light gray
    data_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
    center_align = Alignment(horizontal="center", vertical="center")
    bold_font = Font(bold=True)
    
    # Calculate pass rates
    def calculate_pass_rate(results):
        total = len(results)
        chatgpt_passed = sum(1 for r in results if r.get("chatgpt_validity") == "Valid")
        deepseek_passed = sum(1 for r in results if r.get("deepseek_validity") == "Valid")
        return total, chatgpt_passed, deepseek_passed
    
    airline_total, airline_gpt_passed, airline_ds_passed = calculate_pass_rate(airline_results)
    visa_total, visa_gpt_passed, visa_ds_passed = calculate_pass_rate(visa_results)
    
    # Calculate percentages
    airline_gpt_pct = (airline_gpt_passed / airline_total * 100) if airline_total > 0 else 0
    airline_ds_pct = (airline_ds_passed / airline_total * 100) if airline_total > 0 else 0
    visa_gpt_pct = (visa_gpt_passed / visa_total * 100) if visa_total > 0 else 0
    visa_ds_pct = (visa_ds_passed / visa_total * 100) if visa_total > 0 else 0
    
    # Header row 1
    ws.merge_cells('A1:B1')
    ws['A1'] = "Pass"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_align
    
    ws.merge_cells('C1:D1')
    ws['C1'] = "Emirates Airline Policy"
    ws['C1'].font = header_font
    ws['C1'].fill = header_fill
    ws['C1'].alignment = center_align
    
    ws.merge_cells('E1:F1')
    ws['E1'] = "UAE Visa Guidance"
    ws['E1'].font = header_font
    ws['E1'].fill = header_fill
    ws['E1'].alignment = center_align
    
    # Header row 2
    ws['A2'] = ""
    ws['A2'].fill = header_fill
    ws['B2'] = ""
    ws['B2'].fill = header_fill
    ws['C2'] = "Rate"
    ws['C2'].font = header_font
    ws['C2'].fill = header_fill
    ws['C2'].alignment = center_align
    ws['D2'] = "Percent"
    ws['D2'].font = header_font
    ws['D2'].fill = header_fill
    ws['D2'].alignment = center_align
    ws['E2'] = "Rate"
    ws['E2'].font = header_font
    ws['E2'].fill = header_fill
    ws['E2'].alignment = center_align
    ws['F2'] = "Percent"
    ws['F2'].font = header_font
    ws['F2'].fill = header_fill
    ws['F2'].alignment = center_align
    
    # ChatGPT row
    ws['A3'] = "ChatGPT"
    ws['A3'].font = bold_font
    ws['A3'].fill = data_fill_odd
    ws['B3'] = ""
    ws['B3'].fill = data_fill_odd
    
    ws['C3'] = f"{airline_gpt_passed}/{airline_total}"
    ws['C3'].alignment = center_align
    ws['C3'].fill = data_fill_odd
    
    ws['D3'] = f"{airline_gpt_pct:.0f}%"
    ws['D3'].alignment = center_align
    ws['D3'].fill = data_fill_odd
    
    ws['E3'] = f"{visa_gpt_passed}/{visa_total}"
    ws['E3'].alignment = center_align
    ws['E3'].fill = data_fill_odd
    
    ws['F3'] = f"{visa_gpt_pct:.0f}%"
    ws['F3'].alignment = center_align
    ws['F3'].fill = data_fill_odd
    
    # DeepSeek row
    ws['A4'] = "Deepseek"
    ws['A4'].font = bold_font
    ws['A4'].fill = data_fill_even
    ws['B4'] = ""
    ws['B4'].fill = data_fill_even
    
    ws['C4'] = f"{airline_ds_passed}/{airline_total}"
    ws['C4'].alignment = center_align
    ws['C4'].fill = data_fill_even
    
    ws['D4'] = f"{airline_ds_pct:.0f}%"
    ws['D4'].alignment = center_align
    ws['D4'].fill = data_fill_even
    
    ws['E4'] = f"{visa_ds_passed}/{visa_total}"
    ws['E4'].alignment = center_align
    ws['E4'].fill = data_fill_even
    
    ws['F4'] = f"{visa_ds_pct:.0f}%"
    ws['F4'].alignment = center_align
    ws['F4'].fill = data_fill_even
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Add caption
    ws['A6'] = "Table 1. Test Case Summary of Passing Results"
    ws['A6'].font = Font(italic=True)
    
    try:
        wb.save(filename)
        os.chmod(filename, 0o644)  # Make file readable
    except PermissionError as e:
        raise PermissionError(f"Cannot save Excel file '{filename}': Permission denied. Make sure the file is not open in another program.")
    except Exception as e:
        raise Exception(f"Error saving Excel file '{filename}': {str(e)}")
